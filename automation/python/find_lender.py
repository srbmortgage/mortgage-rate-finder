from __future__ import annotations

import argparse
import json
import re
import tempfile
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_LENDER_WORKBOOK_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1iGgdPOYfdsr4b_omUxZSjz9z0Wq8uZvo/export?format=xlsx"
)


def normalize(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def header_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9/ ]+", "", str(value or "").lower()).strip()


def download(url: str) -> Path:
    target = Path(tempfile.gettempdir()) / "Logins_LenderContactList.xlsx"
    urllib.request.urlretrieve(url, target)
    return target


def first_non_empty(row: tuple[Any, ...], indexes: list[int]) -> str:
    for index in indexes:
        if index < len(row) and row[index]:
            return str(row[index]).strip()
    return ""


def parse_credentials(text: str) -> dict[str, str]:
    if not text:
        return {"username": "", "password": ""}
    lines = [line.strip() for line in re.split(r"[\r\n;]+", text) if line.strip()]
    joined = " ".join(lines)
    username_match = re.search(r"(?:user(?:name|id)?|login|email)\s*[:=-]\s*([^\s,|]+)", joined, re.I)
    password_match = re.search(r"(?:pwd|pass(?:word)?)\s*[:=-]\s*([^\s,|]+)", joined, re.I)
    compact_parts = []
    if len(lines) == 1 and not re.search(r"\s", lines[0]):
        compact_parts = [part.strip() for part in re.split(r"[/|,]", lines[0]) if part.strip()]
    username = (
        username_match.group(1)
        if username_match
        else (compact_parts[0] if len(compact_parts) >= 2 else (lines[0] if len(lines) >= 2 else ""))
    )
    password = (
        password_match.group(1)
        if password_match
        else (compact_parts[1] if len(compact_parts) >= 2 else (lines[1] if len(lines) >= 2 else ""))
    )
    return {"username": username.strip("\"' "), "password": password.strip("\"' ")}


def preferred_sheet_order(sheet_names: list[str], preferred: list[str]) -> list[str]:
    selected = []
    for preferred_name in preferred:
        match = next((name for name in sheet_names if name.lower() == preferred_name.lower()), None)
        if match:
            selected.append(match)
    return selected + [name for name in sheet_names if name not in selected]


def find_lender(workbook_path: Path, lender_name: str, preferred_sheets: list[str]) -> dict[str, Any] | None:
    target = normalize(lender_name)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    candidates: list[dict[str, Any]] = []

    sheets_by_name = {sheet.title: sheet for sheet in workbook.worksheets}
    for sheet_name in preferred_sheet_order(workbook.sheetnames, preferred_sheets):
        sheet = sheets_by_name[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        header = None
        for row in rows:
            keys = [header_key(cell) for cell in row]
            if any("lender name" in key for key in keys):
                header = list(row)
                break
        if not header:
            continue

        keys = [header_key(cell) for cell in header]
        lender_col = next((i for i, key in enumerate(keys) if "lender name" in key), None)
        if lender_col is None:
            continue

        login_url_cols = [
            i
            for i, key in enumerate(keys)
            if "loginurl" in key
        ]
        website_cols = [
            i
            for i, key in enumerate(keys)
            if "wholesale website" in key
        ]
        credential_cols = [
            i
            for i, key in enumerate(keys)
            if "login / pwd" in key or "userid / pwd" in key or "password" in key
        ]

        for row in rows:
            raw_name = row[lender_col] if lender_col < len(row) else ""
            if target not in normalize(raw_name):
                continue
            login_url = first_non_empty(row, login_url_cols) or first_non_empty(row, website_cols)
            credential_text = first_non_empty(row, credential_cols)
            credentials = parse_credentials(credential_text)
            candidates.append(
                {
                    "sheetName": sheet.title,
                    "lenderName": str(raw_name).strip(),
                    "loginUrl": login_url,
                    "username": credentials["username"],
                    "password": credentials["password"],
                    "hasUsername": bool(credentials["username"]),
                    "hasPassword": bool(credentials["password"]),
                    "credentialSource": "workbook" if credential_text else "",
                }
            )

    if not candidates:
        return None
    return next(
        (
            item
            for item in candidates
            if item["loginUrl"] and item["hasUsername"] and item["hasPassword"]
        ),
        next((item for item in candidates if item["loginUrl"]), candidates[0]),
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Find a lender row without printing credential values.")
    parser.add_argument("--lender", default="Pennymac")
    parser.add_argument(
        "--preferred-sheet",
        action="append",
        default=["Conventional"],
        help="Sheet to search first. Can be repeated.",
    )
    parser.add_argument("--url", default=DEFAULT_LENDER_WORKBOOK_URL)
    parser.add_argument("--file", help="Use a local lender workbook instead of downloading.")
    args = parser.parse_args()

    workbook_path = Path(args.file).resolve() if args.file else download(args.url)
    result = find_lender(workbook_path, args.lender, args.preferred_sheet)
    if not result:
        raise SystemExit(f"No lender match found for {args.lender}")
    safe_result = {key: value for key, value in result.items() if key not in {"username", "password"}}
    print(json.dumps(safe_result, indent=2))


if __name__ == "__main__":
    main()
